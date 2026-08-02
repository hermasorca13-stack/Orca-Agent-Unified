"""
Tests for skills/xlsx_skill.py.

Unit tests use a tmp directory and the skill's own public API. No
network. openpyxl is required (it's in requirements.txt).
"""
from __future__ import annotations

import importlib.util
import sys
import zipfile
from pathlib import Path
from unittest.mock import patch

import pytest

ROOT = Path(__file__).resolve().parent.parent
SKILL_PATH = ROOT / "skills" / "xlsx_skill.py"


def _load_skill():
    spec = importlib.util.spec_from_file_location(
        "xlsx_skill_under_test", str(SKILL_PATH)
    )
    mod = importlib.util.module_from_spec(spec)
    sys.modules["xlsx_skill_under_test"] = mod
    spec.loader.exec_module(mod)
    return mod


mod = _load_skill()
info = mod.info
list_sheets = mod.list_sheets
read = mod.read
read_cell = mod.read_cell
create = mod.create
append_rows = mod.append_rows
set_cell = mod.set_cell
format_info = mod.format_info
format_table = mod.format_table
XlsxError = mod.XlsxError


# ----------------------------------------------------------------------
# Fixtures
# ----------------------------------------------------------------------
@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a sample .xlsx with two sheets and headers."""
    p = tmp_path / "sample.xlsx"
    create(
        str(p),
        headers=["Name", "Score", "Active"],
        data=[
            ["Alice", 95, True],
            ["Bob", 87, False],
            ["Carol", 100, True],
        ],
        sheet_name="Users",
        author="Tester",
    )
    # Add a second sheet.
    append_rows_or_create_second_sheet(p, "Notes", [["first note"], ["second note"]])
    return p


def append_rows_or_create_second_sheet(p, name, rows):
    """Helper: open the workbook, add a second sheet, and save."""
    import openpyxl
    wb = openpyxl.load_workbook(str(p))
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    ws = wb[name]
    for r_off, row in enumerate(rows, 1):
        for c_idx, v in enumerate(row, 1):
            ws.cell(row=r_off, column=c_idx, value=v)
    wb.save(str(p))
    wb.close()


# ----------------------------------------------------------------------
# Read
# ----------------------------------------------------------------------
class TestRead:
    def test_info_basic(self, sample_xlsx):
        meta = info(sample_xlsx)
        assert "Users" in meta["sheets"]
        assert "Notes" in meta["sheets"]
        assert meta["sheet_count"] == 2
        assert meta["creator"] == "Tester"
        assert meta["size_bytes"] > 0

    def test_list_sheets(self, sample_xlsx):
        sheets = list_sheets(sample_xlsx)
        assert sheets == ["Users", "Notes"]

    def test_read_default_sheet(self, sample_xlsx):
        r = read(sample_xlsx)
        assert r["sheet_name"] == "Users"
        assert r["headers"] == ["Name", "Score", "Active"]
        assert r["rows"] == [
            ["Alice", 95, True],
            ["Bob", 87, False],
            ["Carol", 100, True],
        ]
        assert r["total_rows"] == 4  # header + 3 data rows
        assert r["truncated"] is False

    def test_read_by_sheet_name(self, sample_xlsx):
        r = read(sample_xlsx, sheet="Notes")
        assert r["sheet_name"] == "Notes"

    def test_read_by_sheet_index(self, sample_xlsx):
        r = read(sample_xlsx, sheet=1)
        assert r["sheet_name"] == "Notes"

    def test_read_max_rows(self, sample_xlsx):
        r = read(sample_xlsx, max_rows=2)
        assert r["truncated"] is True
        assert len(r["rows"]) == 1  # only first data row (header + 1 = 2 rows)

    def test_read_max_cols(self, sample_xlsx):
        r = read(sample_xlsx, max_cols=2)
        # Only Name + Score columns retained.
        assert len(r["headers"]) == 2
        for row in r["rows"]:
            assert len(row) == 2

    def test_read_unknown_sheet(self, sample_xlsx):
        with pytest.raises(XlsxError, match="Sheet not found"):
            read(sample_xlsx, sheet="NoSuch")

    def test_read_sheet_index_out_of_range(self, sample_xlsx):
        with pytest.raises(XlsxError, match="out of range"):
            read(sample_xlsx, sheet=99)

    def test_read_cell(self, sample_xlsx):
        assert read_cell(sample_xlsx, "Users", "A1") == "Name"
        assert read_cell(sample_xlsx, "Users", "A2") == "Alice"
        assert read_cell(sample_xlsx, "Users", "B2") == 95
        assert read_cell(sample_xlsx, 0, "C4") is True  # Carol's Active

    def test_read_cell_invalid_ref(self, sample_xlsx):
        with pytest.raises(XlsxError, match="Invalid cell reference"):
            read_cell(sample_xlsx, "Users", "ZZ")


# ----------------------------------------------------------------------
# Write
# ----------------------------------------------------------------------
class TestWrite:
    def test_create_minimal(self, tmp_path):
        p = create(str(tmp_path / "min.xlsx"))
        assert Path(p).exists()
        assert Path(p).suffix == ".xlsx"

    def test_create_with_data(self, tmp_path):
        p = create(
            str(tmp_path / "data.xlsx"),
            headers=["a", "b"],
            data=[[1, 2], [3, 4]],
        )
        r = read(p)
        assert r["headers"] == ["a", "b"]
        assert r["rows"] == [[1, 2], [3, 4]]

    def test_create_appends_xlsx_extension(self, tmp_path):
        p = create(str(tmp_path / "noext"))
        assert Path(p).suffix == ".xlsx"

    def test_append_rows(self, tmp_path):
        p = create(
            str(tmp_path / "a.xlsx"),
            headers=["x"],
            data=[[1], [2]],
        )
        added = append_rows(p, "Sheet1", [[3], [4], [5]])
        assert added == 3
        r = read(p)
        assert r["rows"] == [[1], [2], [3], [4], [5]]

    def test_append_unknown_sheet(self, sample_xlsx):
        with pytest.raises(XlsxError, match="Sheet not found"):
            append_rows(sample_xlsx, "NoSuch", [[1]])

    def test_append_empty_raises(self, sample_xlsx):
        with pytest.raises(XlsxError, match="No rows to append"):
            append_rows(sample_xlsx, "Users", [])

    def test_set_cell(self, tmp_path):
        p = create(
            str(tmp_path / "s.xlsx"),
            headers=["a", "b"],
            data=[[1, 2]],
        )
        set_cell(p, "Sheet1", "B2", 999)
        assert read_cell(p, "Sheet1", "B2") == 999

    def test_set_cell_invalid_ref(self, tmp_path):
        p = create(str(tmp_path / "s.xlsx"))
        with pytest.raises(XlsxError, match="Invalid cell reference"):
            set_cell(p, "Sheet1", "1A", 5)


# ----------------------------------------------------------------------
# Errors
# ----------------------------------------------------------------------
class TestErrors:
    def test_missing_file(self):
        with pytest.raises(XlsxError, match="File not found"):
            info("/no/such/file.xlsx")

    def test_legacy_xls(self, tmp_path):
        p = tmp_path / "old.xls"
        p.write_bytes(b"old")
        with pytest.raises(XlsxError, match="Legacy"):
            info(p)

    def test_bad_extension(self, tmp_path):
        bad = tmp_path / "x.txt"
        bad.write_text("hi")
        with pytest.raises(XlsxError, match="Unsupported extension"):
            info(bad)

    def test_not_a_zip(self, tmp_path):
        fake = tmp_path / "fake.xlsx"
        fake.write_bytes(b"not a zip")
        with pytest.raises(XlsxError, match="not a zip"):
            info(fake)

    def test_zip_missing_content_types(self, tmp_path):
        p = tmp_path / "weird.xlsx"
        with zipfile.ZipFile(p, "w") as z:
            z.writestr("hello.txt", "hi")
        with pytest.raises(XlsxError, match="Content_Types"):
            info(p)

    def test_oversize_rejected(self, tmp_path):
        with patch.object(mod, "MAX_BYTES", 4):
            p = tmp_path / "ok.xlsx"
            p.write_bytes(b"abcde")
            with pytest.raises(XlsxError, match="too large"):
                info(p)


# ----------------------------------------------------------------------
# Format helpers
# ----------------------------------------------------------------------
class TestFormatters:
    def test_format_info(self):
        meta = {
            "title": "Sales",
            "sheets": ["Q1", "Q2"],
            "active_sheet": "Q1",
            "total_cells": 100,
            "creator": "Me",
            "modified": "2026-01-15T10:00:00",
            "size_bytes": 1234,
        }
        out = format_info(meta)
        assert "Sales" in out
        assert "Q1" in out and "Q2" in out
        assert "1,234 bytes" in out

    def test_format_table_empty(self):
        out = format_table([], [], sheet_name="Empty")
        assert "empty" in out.lower()

    def test_format_table_with_data(self):
        out = format_table(
            ["a", "b"],
            [["1", "2"], ["3", "4"]],
            sheet_name="S",
        )
        assert "Sheet: S" in out
        assert "a | b" in out
        assert "1 | 2" in out

    def test_format_table_escapes_pipes(self):
        out = format_table(["col"], [["a|b"]])
        # Pipes in cell values must be escaped so the table renders right.
        assert "a\\|b" in out

    def test_format_table_truncates(self):
        rows = [[i] for i in range(100)]
        out = format_table(["n"], rows, max_rows=5)
        assert "95 more rows" in out


class TestSupportedExts:
    def test_xlsx_and_xlsm(self):
        assert ".xlsx" in mod.SUPPORTED_EXTS
        assert ".xlsm" in mod.SUPPORTED_EXTS
