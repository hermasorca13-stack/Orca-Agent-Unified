"""Sensory Perception Module for Orca Agent"""

import os
import asyncio
from typing import Dict, Any, List
from loguru import logger

# External libraries for file processing (with graceful fallback)
import pypdf
import openpyxl
import docx
from PIL import Image
import pytesseract

# Heavy / optional deps — guarded so the module can still import on minimal envs
try:
    import easyocr  # type: ignore
    _EASYOCR_AVAILABLE = True
except Exception:  # ImportError or backend failure
    easyocr = None
    _EASYOCR_AVAILABLE = False

try:
    import cv2  # type: ignore
    _CV2_AVAILABLE = True
except Exception:
    cv2 = None
    _CV2_AVAILABLE = False

try:
    import numpy as np  # type: ignore
    _NUMPY_AVAILABLE = True
except Exception:
    np = None
    _NUMPY_AVAILABLE = False

# Manus tools integration (simulated for now, actual calls will be via shell)
async def manus_analyze_video_mock(video_url_or_path: str, prompt: str) -> str:
    logger.info(f"[Manus Mock] Analyzing video: {video_url_or_path} with prompt: {prompt}")
    await asyncio.sleep(1) # Simulate async operation
    return f"Video analysis result for {video_url_or_path}: Key frames detected, objects identified based on '{prompt}'."

async def manus_speech_to_text_mock(input_file: str) -> str:
    logger.info(f"[Manus Mock] Transcribing audio: {input_file}")
    await asyncio.sleep(0.5) # Simulate async operation
    return f"Audio transcription result for {input_file}: 'This is a sample transcription.'"


class SensoryPerception:
    """Handles all sensory input processing for the Orca Agent"""

    def __init__(self):
        logger.info("👁️ Initializing Sensory Perception module...")
        # Lazy init EasyOCR — only build the reader if the dep is available
        # and the user actually requests OCR. Keeps import cheap.
        self.ocr_reader = None
        if _EASYOCR_AVAILABLE:
            try:
                self.ocr_reader = easyocr.Reader(["en", "ar"], gpu=False, verbose=False)
            except Exception as e:
                logger.warning(f"⚠️ EasyOCR init failed, will use Tesseract only: {e}")
        else:
            logger.warning("⚠️ easyocr not installed — falling back to Tesseract only")

        self.tesseract_path = os.getenv("TESSERACT_PATH", "/usr/bin/tesseract")
        try:
            pytesseract.pytesseract.tesseract_cmd = self.tesseract_path
        except Exception:
            pass
        logger.info("✅ Sensory Perception module initialized.")

    async def process_file(self, file_path: str, file_type: str, options: Dict = None) -> Dict:
        """Process various file types and extract information"""
        logger.info(f"Processing file: {file_path} (Type: {file_type})")
        options = options or {}
        
        try:
            if file_type == "pdf":
                return await self._process_pdf(file_path, options)
            elif file_type == "excel":
                return await self._process_excel(file_path, options)
            elif file_type == "word":
                return await self._process_word(file_path, options)
            elif file_type == "image":
                return await self._process_image(file_path, options)
            elif file_type == "video":
                return await self._process_video(file_path, options)
            elif file_type == "audio":
                return await self._process_audio(file_path, options)
            elif file_type == "3d_model":
                return await self._process_3d_model(file_path, options)
            elif file_type == "cad":
                return await self._process_cad(file_path, options)
            elif file_type == "legacy_code":
                return await self._process_legacy_code(file_path, options)
            else:
                logger.warning(f"Unsupported file type: {file_type}")
                return {"status": "failed", "message": "Unsupported file type"}
        except Exception as e:
            logger.error(f"Error processing {file_type} file {file_path}: {e}")
            return {"status": "error", "message": str(e)}

    async def _process_pdf(self, file_path: str, options: Dict) -> Dict:
        """Extract text from PDF files"""
        text_content = []
        try:
            with open(file_path, "rb") as f:
                reader = pypdf.PdfReader(f)
                for page in reader.pages:
                    text_content.append(page.extract_text())
            return {"status": "success", "content": "\n".join(text_content)}
        except Exception as e:
            return {"status": "error", "message": f"PDF processing failed: {e}"}

    async def _process_excel(self, file_path: str, options: Dict) -> Dict:
        """Extract data from Excel files"""
        data = {}
        try:
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(list(row))
                data[sheet_name] = sheet_data
            return {"status": "success", "content": data}
        except Exception as e:
            return {"status": "error", "message": f"Excel processing failed: {e}"}

    async def _process_word(self, file_path: str, options: Dict) -> Dict:
        """Extract text from Word documents"""
        text_content = []
        try:
            document = docx.Document(file_path)
            for paragraph in document.paragraphs:
                text_content.append(paragraph.text)
            return {"status": "success", "content": "\n".join(text_content)}
        except Exception as e:
            return {"status": "error", "message": f"Word processing failed: {e}"}

    async def _process_image(self, file_path: str, options: Dict) -> Dict:
        """Extract text from images (OCR) and analyze content"""
        text_easyocr = ""
        if self.ocr_reader is not None:
            try:
                ocr_result_easyocr = self.ocr_reader.readtext(file_path, detail=0)
                text_easyocr = " ".join(ocr_result_easyocr)
            except Exception as e:
                logger.warning(f"EasyOCR failed: {e}")
        else:
            text_easyocr = "[EasyOCR unavailable on this host]"

        text_tesseract = ""
        try:
            image = Image.open(file_path)
            text_tesseract = pytesseract.image_to_string(image, lang=options.get("lang", "eng"))
        except Exception as e:
            text_tesseract = f"[Tesseract failed: {e}]"

        analysis_result = await self._analyze_image_content(file_path, options)

        return {
            "status": "success",
            "text_easyocr": text_easyocr,
            "text_tesseract": text_tesseract,
            "analysis": analysis_result,
        }

    async def _analyze_image_content(self, image_path: str, options: Dict) -> str:
        """Mock for advanced image content analysis (e.g., object detection, scene understanding)"""
        logger.info(f"[Mock] Analyzing image content for {image_path}")
        await asyncio.sleep(0.3) # Simulate processing
        return "Detected common objects and scene elements. This is a mock analysis."

    async def _process_video(self, file_path: str, options: Dict) -> Dict:
        """Analyze video content frame-by-frame and extract key information"""
        prompt = options.get("prompt", "summarize key events and objects")
        try:
            # Use Manus tool for video analysis
            analysis_result = await manus_analyze_video_mock(file_path, prompt)
            return {"status": "success", "analysis": analysis_result}
        except Exception as e:
            return {"status": "error", "message": f"Video processing failed: {e}"}

    async def _process_audio(self, file_path: str, options: Dict) -> Dict:
        """Transcribe audio files (voice notes) to text"""
        try:
            # Use Manus tool for speech-to-text
            transcription = await manus_speech_to_text_mock(file_path)
            return {"status": "success", "transcription": transcription}
        except Exception as e:
            return {"status": "error", "message": f"Audio processing failed: {e}"}

    async def _process_3d_model(self, file_path: str, options: Dict) -> Dict:
        """Process 3D model files (placeholder for future integration)"""
        logger.warning(f"3D model processing not fully implemented: {file_path}")
        return {"status": "pending", "message": "3D model processing requires specialized libraries and integration."}

    async def _process_cad(self, file_path: str, options: Dict) -> Dict:
        """Process CAD files (placeholder for future integration)"""
        logger.warning(f"CAD file processing not fully implemented: {file_path}")
        return {"status": "pending", "message": "CAD file processing requires specialized libraries and integration."}

    async def _process_legacy_code(self, file_path: str, options: Dict) -> Dict:
        """Read and analyze legacy code files"""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                code_content = f.read()
            # Placeholder for actual code analysis (e.g., dependency parsing, complexity analysis)
            analysis = {"lines_of_code": len(code_content.splitlines()), "file_size": os.path.getsize(file_path)}
            return {"status": "success", "content": code_content, "analysis": analysis}
        except Exception as e:
            return {"status": "error", "message": f"Legacy code processing failed: {e}"}

    async def read_screen_content(self, screenshot_path: str, options: Dict = None) -> Dict:
        """Read and interpret screen content from screenshots"""
        options = options or {}
        try:
            # Use image processing for OCR on screenshots
            result = await self._process_image(screenshot_path, options)
            result["source"] = "screenshot"
            return result
        except Exception as e:
            return {"status": "error", "message": f"Screen content reading failed: {e}"}

    async def analyze_sensor_stream(self, stream_data: List[Dict], options: Dict = None) -> Dict:
        """Analyze real-time sensor stream data (IoT-style)"""
        options = options or {}
        logger.info(f"[Mock] Analyzing sensor stream with {len(stream_data)} data points.")
        await asyncio.sleep(0.5) # Simulate processing
        # Placeholder for actual sensor data analysis (e.g., anomaly detection, pattern recognition)
        analysis_result = {"summary": "Simulated analysis of sensor data. Detected normal operation.", "data_points": len(stream_data)}
        return {"status": "success", "analysis": analysis_result}

    async def ocr_blurry_image(self, image_path: str, options: Dict = None) -> Dict:
        """Perform OCR on blurry images, attempting to enhance first"""
        options = options or {}
        try:
            # Simulate image enhancement (e.g., deblurring, contrast adjustment)
            logger.info(f"[Mock] Enhancing blurry image: {image_path}")
            # Then perform OCR
            result = await self._process_image(image_path, options)
            result["enhancement_note"] = "Image enhancement simulated before OCR."
            return result
        except Exception as e:
            return {"status": "error", "message": f"Blurry image OCR failed: {e}"}

    async def understand_blueprints(self, blueprint_image_path: str, options: Dict = None) -> Dict:
        """Understand and interpret engineering blueprints (placeholder)"""
        logger.warning(f"Blueprint understanding not fully implemented: {blueprint_image_path}")
        return {"status": "pending", "message": "Blueprint understanding requires specialized computer vision and domain knowledge."}


# Example usage (for testing purposes)
async def main():
    sensory = SensoryPerception()
    
    # Mock files for testing
    # Create dummy files for testing
    with open("dummy.pdf", "w") as f: f.write("This is a dummy PDF content.")
    with open("dummy.xlsx", "w") as f: f.write("This is a dummy Excel content.")
    with open("dummy.docx", "w") as f: f.write("This is a dummy Word content.")
    Image.new("RGB", (100, 100), color = 'red').save("dummy.png")
    with open("dummy.mp3", "w") as f: f.write("dummy audio data")
    with open("dummy.mp4", "w") as f: f.write("dummy video data")
    with open("dummy.py", "w") as f: f.write("print('Hello, World!')")

    print("\n--- Processing PDF ---")
    pdf_result = await sensory.process_file("dummy.pdf", "pdf")
    print(pdf_result)

    print("\n--- Processing Excel ---")
    excel_result = await sensory.process_file("dummy.xlsx", "excel")
    print(excel_result)

    print("\n--- Processing Word ---")
    word_result = await sensory.process_file("dummy.docx", "word")
    print(word_result)

    print("\n--- Processing Image (OCR) ---")
    image_result = await sensory.process_file("dummy.png", "image")
    print(image_result)

    print("\n--- Processing Video ---")
    video_result = await sensory.process_file("dummy.mp4", "video", {"prompt": "detect motion"})
    print(video_result)

    print("\n--- Processing Audio ---")
    audio_result = await sensory.process_file("dummy.mp3", "audio")
    print(audio_result)

    print("\n--- Processing Legacy Code ---")
    code_result = await sensory.process_file("dummy.py", "legacy_code")
    print(code_result)

    print("\n--- Reading Screen Content (Mock) ---")
    screen_result = await sensory.read_screen_content("dummy.png")
    print(screen_result)

    print("\n--- Analyzing Sensor Stream (Mock) ---")
    sensor_data = [{"temp": 25, "humidity": 60}, {"temp": 26, "humidity": 61}]
    sensor_result = await sensory.analyze_sensor_stream(sensor_data)
    print(sensor_result)

    print("\n--- OCR Blurry Image (Mock) ---")
    blurry_ocr_result = await sensory.ocr_blurry_image("dummy.png")
    print(blurry_ocr_result)

    # Clean up dummy files
    os.remove("dummy.pdf")
    os.remove("dummy.xlsx")
    os.remove("dummy.docx")
    os.remove("dummy.png")
    os.remove("dummy.mp3")
    os.remove("dummy.mp4")
    os.remove("dummy.py")

if __name__ == "__main__":
    asyncio.run(main())
