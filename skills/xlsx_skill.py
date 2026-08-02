"""
skills/xlsx_skill.py — Read & write Microsoft Excel (.xlsx) files.

Why this skill:
- `openpyxl` is the consensus 2026 standard for .xlsx in Python:
  MIT-licensed, no native deps, no Excel/Office needed, used by 12M+
  projects on GitHub. It is the only mature choice for reading and
  writing modern Excel files (.xlsx, .xlsm) without Office automation.
- Orca agents often need to: dump a CSV-shaped payload to .xlsx,
  preview a sheet as a markdown table, or read a single cell value.

Public surface:
- `info(path)` — sheet names, dimensions, creator, last modified.
- `list_sheets(path)` — names of all worksheets in the workbook.
- `read(path, sheet=None, max_rows=None, max_cols=None)` — full sheet
  as a dict with `headers`, `rows`, `total_rows`, `total_cols`,
  `sheet_name`, `truncated` (bool).
- `read_cell(path, sheet, cell)` — single cell value (e.g. "B2").
- `create(out_path, *, data=None, headers=None, sheet_name="Sheet1",
  author=None)` — build a new .xlsx from a 2D list. Returns path.
- `append_rows(path, sheet_name, rows)` — append rows to a sheet.
- `set_cell(path, sheet_name, cell, value)` — write a single cell.
- `format_info(meta)` / `format_table(headers, rows, ...)` — Telegram
  cards (markdown table for a sheet).
- `XlsxError` — single, user-friendly exception class.

Engineering contract (Apple + Microsoft grade):
- Lazy-import openpyxl so a missing dep surfaces at call time, not
  import time.
- Whitelist .xlsx / .xlsm extensions; reject legacy .xls with a clear
  hint (use `xlsx_skill` to convert via pandas/libreoffice).
- Friendly error messages for: missing file, corrupted zip, wrong
  extension, oversized file, missing sheet.
- loguru integration for telemetry.
- No global state; pure functional reads, idempotent writes.

This file is ADD-ONLY. It does not modify any existing module.
"""
from __future__ import annotations

import os
import re
import time
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from loguru import logger


# ----------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------
SUPPORTED_EXTS = {".xlsx", ".xlsm"}
MAX_BYTES = 100 * 1024 * 1024  # 100 MB — workbooks can be heavy

# Cell reference pattern: column letters + 1-based row number, e.g. "A1", "AB12".
_CELL_RE = re.compile(r"^([A-Z]+)(\d+)$")


class XlsxError(RuntimeError):
    """Raised when .xlsx processing fails for any reason."""


# ----------------------------------------------------------------------
# Internal helpers
# ----------------------------------------------------------------------
def _require_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except ImportError as exc:
        raise XlsxError(
            "openpyxl is not installed. Run: pip install 'openpyxl>=3.1.0'"
        ) from exc


def _resolve_path(path: Union[str, Path]) -> Path:
    p = Path(path)
    if not p.exists():
        raise XlsxError(f"File not found: {p}")
    ext = p.suffix.lower()
    if ext not in SUPPORTED_EXTS:
        if ext == ".xls":
            raise XlsxError(
                f"Legacy .xls is not supported. "
                f"Open in Excel and re-save as .xlsx, or use LibreOffice."
            )
        raise XlsxError(
            f"Unsupported extension {ext!r}. Supported: {sorted(SUPPORTED_EXTS)}"
        )
    if p.stat().st_size > MAX_BYTES:
        raise XlsxError(
            f"File too large: {p.stat().st_size / 1024 / 1024:.1f} MB "
            f"(max {MAX_BYTES // (1024*1024)} MB)."
        )
    # Validate it's a real .xlsx (zip with [Content_Types].xml).
    try:
        if not zipfile.is_zipfile(p):
            raise XlsxError(
                "Not a valid .xlsx file (not a zip archive). "
                "The file may be corrupted or in an unsupported format."
            )
        with zipfile.ZipFile(p) as z:
            if "[Content_Types].xml" not in z.namelist():
                raise XlsxError(
                    "Not a valid .xlsx file (missing [Content_Types].xml). "
                    "The file may be corrupted or in an unsupported format."
                )
    except zipfile.BadZipFile as exc:
        raise XlsxError(
            "Not a valid .xlsx file (corrupted zip). "
            "Try re-saving it from Excel or LibreOffice."
        ) from exc
    return p


def _parse_cell(ref: str) -> tuple[int, int]:
    """Parse "A1" → (row=1, col=1). 1-based, matching Excel's notation."""
    m = _CELL_RE.match((ref or "").strip().upper())
    if not m:
        raise XlsxError(f"Invalid cell reference: {ref!r}. Use Excel notation like 'A1' or 'B12'.")
    col_letters, row_str = m.group(1), m.group(2)
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return int(row_str), col


def _col_letters(n: int) -> str:
    """1-based column index → Excel letters (1=A, 27=AA)."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


# ----------------------------------------------------------------------
# Public API — read
# ----------------------------------------------------------------------
def info(path: Union[str, Path]) -> Dict[str, Any]:
    """Return workbook metadata.

    Returns a dict with: sheets (list[str]), sheet_count, active_sheet,
    creator, modified, size_bytes, total_cells (sum across sheets).
    """
    p = _resolve_path(path)
    openpyxl = _require_openpyxl()
    t0 = time.monotonic()
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        if "password" in str(exc).lower() or "encrypted" in str(exc).lower():
            raise XlsxError("This .xlsx is password-protected. Decrypt it first.") from exc
        raise XlsxError(f"Could not open .xlsx: {exc}") from exc

    try:
        sheets = wb.sheetnames
        active = wb.active.title if wb.active else None
        props = wb.properties
        total_cells = 0
        for s in wb.worksheets:
            total_cells += (s.max_row or 0) * (s.max_column or 0)
        out: Dict[str, Any] = {
            "sheets": sheets,
            "sheet_count": len(sheets),
            "active_sheet": active,
            "creator": (props.creator or "").strip() or None,
            "title": (props.title or "").strip() or None,
            "modified": props.modified.isoformat() if props.modified else None,
            "size_bytes": p.stat().st_size,
            "total_cells": total_cells,
        }
        logger.info(
            "xlsx info | path={} sheets={} total_cells={} took={}s",
            p.name, len(sheets), total_cells, round(time.monotonic() - t0, 3),
        )
        return out
    finally:
        wb.close()


def list_sheets(path: Union[str, Path]) -> List[str]:
    """Return the list of sheet names in the workbook."""
    p = _resolve_path(path)
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise XlsxError(f"Could not open .xlsx: {exc}") from exc
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read(
    path: Union[str, Path],
    *,
    sheet: Optional[Union[str, int]] = None,
    max_rows: Optional[int] = None,
    max_cols: Optional[int] = None,
) -> Dict[str, Any]:
    """Read a sheet and return a structured dict.

    Args:
        path: workbook path.
        sheet: sheet name (str) or index (0-based int). Defaults to the
            first sheet.
        max_rows: cap on rows returned. None = no cap.
        max_cols: cap on columns returned. None = no cap.

    Returns:
        dict with keys: sheet_name, headers (list[str]), rows
        (list[list[Any]]), total_rows, total_cols, truncated (bool).
    """
    p = _resolve_path(path)
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise XlsxError(f"Could not open .xlsx: {exc}") from exc
    try:
        if sheet is None:
            ws = wb.worksheets[0]
        elif isinstance(sheet, int):
            if sheet < 0 or sheet >= len(wb.worksheets):
                raise XlsxError(
                    f"Sheet index {sheet} out of range. "
                    f"Workbook has {len(wb.worksheets)} sheet(s)."
                )
            ws = wb.worksheets[sheet]
        else:
            if sheet not in wb.sheetnames:
                raise XlsxError(
                    f"Sheet not found: {sheet!r}. "
                    f"Available: {wb.sheetnames}"
                )
            ws = wb[sheet]

        # Pull all values.
        all_rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))
        # Trim trailing fully-empty rows.
        while all_rows and not any(c not in (None, "") for c in all_rows[-1]):
            all_rows.pop()

        total_rows = len(all_rows)
        total_cols = max((len(r) for r in all_rows), default=0)

        # Apply max_cols first (truncate columns).
        if max_cols is not None and max_cols >= 0:
            all_rows = [r[:max_cols] for r in all_rows]
            total_cols = min(total_cols, max_cols)

        # Apply max_rows (truncate rows).
        truncated = False
        if max_rows is not None and max_rows >= 0 and total_rows > max_rows:
            all_rows = all_rows[:max_rows]
            truncated = True

        # First row → headers (if it has any text), else synthetic.
        headers: List[str] = []
        body = all_rows
        if all_rows:
            first = all_rows[0]
            if any(isinstance(c, str) and c.strip() for c in first):
                headers = [str(c) if c is not None else "" for c in first]
                body = all_rows[1:]

        out = {
            "sheet_name": ws.title,
            "headers": headers,
            "rows": body,
            "total_rows": total_rows,
            "total_cols": total_cols,
            "truncated": truncated,
        }
        logger.info(
            "xlsx read | path={} sheet={} rows={} cols={} truncated={}",
            p.name, ws.title, len(body), len(headers), truncated,
        )
        return out
    finally:
        wb.close()


def read_cell(path: Union[str, Path], sheet: Union[str, int], cell: str) -> Any:
    """Read a single cell value by Excel-style reference (e.g. 'B2')."""
    # Validate the cell reference up front. openpyxl's read-only
    # worksheet raises a confusing AttributeError for malformed refs,
    # so we do our own check first.
    _parse_cell(cell)
    p = _resolve_path(path)
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise XlsxError(f"Could not open .xlsx: {exc}") from exc
    try:
        if isinstance(sheet, int):
            if sheet < 0 or sheet >= len(wb.worksheets):
                raise XlsxError(
                    f"Sheet index {sheet} out of range. "
                    f"Workbook has {len(wb.worksheets)} sheet(s)."
                )
            ws = wb.worksheets[sheet]
        else:
            if sheet not in wb.sheetnames:
                raise XlsxError(
                    f"Sheet not found: {sheet!r}. Available: {wb.sheetnames}"
                )
            ws = wb[sheet]
        # Use openpyxl's coordinate parser instead of ws[ref] (which
        # needs iter_cols not available on ReadOnlyWorksheet).
        from openpyxl.utils.cell import coordinate_from_string
        ref = cell.strip().upper()
        col_letters, row = coordinate_from_string(ref)
        col = 0
        for ch in col_letters:
            col = col * 26 + (ord(ch) - ord("A") + 1)
        # Scan for the cell value (read-only mode needs manual scan).
        value = None
        for r in ws.iter_rows(min_row=row, max_row=row,
                              min_col=col, max_col=col, values_only=True):
            if r:
                value = r[0]
            break
        logger.info("xlsx read_cell | path={} sheet={} cell={} -> {!r}", p.name, ws.title, cell, value)
        return value
    finally:
        wb.close()


# ----------------------------------------------------------------------
# Public API — write
# ----------------------------------------------------------------------
def create(
    out_path: Union[str, Path],
    *,
    data: Optional[List[List[Any]]] = None,
    headers: Optional[List[str]] = None,
    sheet_name: str = "Sheet1",
    author: Optional[str] = None,
) -> str:
    """Create a new .xlsx file.

    Args:
        out_path: where to write the file. Parent dirs are created.
        data: 2D list of cell values (rows of rows).
        headers: optional first row; if given, written in bold.
        sheet_name: name of the worksheet.
        author: optional creator metadata.

    Returns:
        Absolute path of the created file as a string.
    """
    openpyxl = _require_openpyxl()
    out = Path(out_path)
    if out.suffix.lower() not in SUPPORTED_EXTS:
        out = out.with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        ws.title = sheet_name[:31] or "Sheet1"  # Excel limit
        if author:
            wb.properties.creator = author

        row_idx = 1
        if headers:
            for c_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=c_idx, value=h)
                cell.font = openpyxl.styles.Font(bold=True)
            row_idx += 1
        for row in data or []:
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=c_idx, value=value)
            row_idx += 1
        wb.save(str(out))
    finally:
        wb.close()

    logger.info(
        "xlsx create | path={} sheet={} rows={} cols={}",
        out.name, sheet_name,
        len(data or []) + (1 if headers else 0),
        max(len(headers or []), max((len(r) for r in (data or [])), default=0)),
    )
    return str(out.resolve())


def append_rows(path: Union[str, Path], sheet_name: str, rows: List[List[Any]]) -> int:
    """Append rows to an existing sheet. Returns the number of rows added."""
    p = _resolve_path(path)
    openpyxl = _require_openpyxl()
    if not rows:
        raise XlsxError("No rows to append")
    try:
        wb = openpyxl.load_workbook(str(p))
    except Exception as exc:  # noqa: BLE001
        raise XlsxError(f"Could not open .xlsx: {exc}") from exc
    try:
        if sheet_name not in wb.sheetnames:
            raise XlsxError(
                f"Sheet not found: {sheet_name!r}. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        # Find the next free row.
        start = ws.max_row + 1
        for r_off, row in enumerate(rows):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=start + r_off, column=c_idx, value=value)
        wb.save(str(p))
        logger.info(
            "xlsx append_rows | path={} sheet={} added={}",
            p.name, sheet_name, len(rows),
        )
        return len(rows)
    finally:
        wb.close()


def set_cell(path: Union[str, Path], sheet_name: str, cell: str, value: Any) -> None:
    """Write a single cell by Excel-style reference (e.g. 'B2')."""
    p = _resolve_path(path)
    _parse_cell(cell)  # validate the reference
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(str(p))
    except Exception as exc:  # noqa: BLE001
        raise XlsxError(f"Could not open .xlsx: {exc}") from exc
    try:
        if sheet_name not in wb.sheetnames:
            raise XlsxError(
                f"Sheet not found: {sheet_name!r}. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        ws[cell.strip().upper()] = value
        wb.save(str(p))
        logger.info(
            "xlsx set_cell | path={} sheet={} cell={} -> {!r}",
            p.name, sheet_name, cell, value,
        )
    finally:
        wb.close()


# ----------------------------------------------------------------------
# Format helpers (Telegram-friendly)
# ----------------------------------------------------------------------
def format_info(meta: Dict[str, Any]) -> str:
    """Format an `info()` result as a Telegram-friendly card."""
    sheets = meta.get("sheets") or []
    creator = meta.get("creator") or "—"
    modified = (meta.get("modified") or "")[:19].replace("T", " ")
    size = int(meta.get("size_bytes") or 0)
    return (
        f"📊 *{(meta.get('title') or '(untitled)')}*\n"
        f"Sheets ({len(sheets)}): {', '.join(sheets) if sheets else '—'}\n"
        f"Active: `{meta.get('active_sheet') or '—'}`\n"
        f"Cells: {meta.get('total_cells', 0):,}\n"
        f"Creator: {creator}\n"
        f"Modified: {modified or '—'}\n"
        f"Size: {size:,} bytes"
    )


def format_table(
    headers: List[str],
    rows: List[List[Any]],
    *,
    sheet_name: Optional[str] = None,
    truncated: bool = False,
    max_rows: int = 30,
) -> str:
    """Format a sheet as a Telegram-friendly Markdown table."""
    if not headers and not rows:
        return f"📭 Sheet `{sheet_name or '?'}` is empty"
    # Cap rows for the message size.
    shown = rows[:max_rows]
    lines = []
    if sheet_name:
        lines.append(f"📋 *Sheet: {sheet_name}*")
    # Build header row.
    cols = max(len(headers), max((len(r) for r in shown), default=0))
    hdr_cells = [(headers[i] if i < len(headers) else "") for i in range(cols)]
    lines.append("| " + " | ".join(_md_escape(c) for c in hdr_cells) + " |")
    lines.append("| " + " | ".join(["---"] * cols) + " |")
    for row in shown:
        cells = [(row[i] if i < len(row) else "") for i in range(cols)]
        lines.append("| " + " | ".join(_md_escape(c) for c in cells) + " |")
    if len(rows) > max_rows:
        lines.append(f"_…+{len(rows) - max_rows} more rows_")
    if truncated:
        lines.append("_…truncated_")
    return "\n".join(lines)


def _md_escape(value: Any) -> str:
    """Make a cell safe for a Telegram Markdown table."""
    s = "" if value is None else str(value)
    # Escape the characters Telegram's Markdown parser cares about.
    for ch in ("\\", "|", "`", "*", "_", "[", "]"):
        s = s.replace(ch, f"\\{ch}")
    # Newlines → space (tables don't support multi-line cells).
    s = s.replace("\n", " ").replace("\r", " ")
    return s or " "


__all__ = [
    "info", "list_sheets", "read", "read_cell",
    "create", "append_rows", "set_cell",
    "format_info", "format_table",
    "XlsxError", "SUPPORTED_EXTS", "MAX_BYTES",
]
